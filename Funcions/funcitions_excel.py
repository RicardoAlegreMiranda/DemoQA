import os
import openpyxl


# Obtén la ruta del directorio actual, para que el acceso al Excel sea con una ruta Relativa en lugar de absoluta
ruta_actual = os.path.dirname(os.path.abspath(__file__))
ruta_modificada = os.path.dirname(ruta_actual)

# Construir la nueva ruta con la parte modificada
ruta_modificada = os.path.join(ruta_modificada, "Test_Dates", "TestDates.xlsx")


class FuncionesExcel:

    def __init__(self, ruta_excel, hoja_nombre):
        self.ruta_excel = ruta_modificada
        self.libro_excel = openpyxl.load_workbook(self.ruta_excel)
        self.hoja_excel = self.libro_excel[hoja_nombre]

    def obtener_num_filas(self):
        return self.hoja_excel.max_row

    def obtener_num_columnas(self):
        return self.hoja_excel.max_column

    def leer_celda(self, fila, columna):
        return self.hoja_excel.cell(row=fila, column=columna).value

    def escribir_celda(self, fila, columna, valor):
        self.hoja_excel.cell(row=fila, column=columna).value = valor
        self.libro_excel.save(self.ruta_excel)

    def cerrar_excel(self):
        self.libro_excel.close()

